import openpyxl
wb1 = openpyxl.load_workbook('C:\\Users\\Administrator\\Desktop\\hehe.xlsx')
wb2 = openpyxl.load_workbook('C:\\Users\\Administrator\\Desktop\\2017精益眼镜奖金表（最终表）.xlsx')
sheet1 = wb1.get_sheet_by_name('大连')
sheet2 = wb2.get_sheet_by_name('大连')
for i in range(6,31,2):
    sheet1.cell(row=9, column=i).value = str(round(float(sheet2.cell(row=i /2 + 1, column=1).value), 1)) + "=" + str(round(int(sheet2.cell(row=i / 2 + 1, column=7).value), 1)) + "*" + str(round(float(sheet2.cell(row=i / 2 + 1, column=6).value), 2))
    sheet1.cell(row=8, column=i).value = str(round(float(sheet2.cell(row=i /2 + 1,column=3).value),1)) + "=" + str(round(int(sheet2.cell(row=i / 2 + 1,column=7).value),1))
wb1.save('C:\\Users\\Administrator\\Desktop\\hehe.xlsx')
